"""
Cross-check: Dashboard API  vs  raw source (PROD.xlsx = DB).
-----------------------------------------------------------
PROD.xlsx hi DB ka source hain. Ye script Excel se SEEDHA (sir wali
methodology se) har din / har shift / har ghante ke Old/New/Test Cell
engines compute karta hain, phir API se compare karta hain.

Sab PASS aaye = jab real DB se connect karenge to numbers exactly sahi rahenge.

Run:  python verify_dashboard.py
"""

import json
import urllib.request
from collections import defaultdict
from datetime import timedelta

import openpyxl

EXCEL = r"C:\Users\Sarfaraz\Downloads\PROD.xlsx"
API = "http://localhost:5000/api/Dashboard/overview"

OLD_LINE = "23800"   # Old Line
NEW_LINE = "33200"   # New Line
TEST_CELL = "TEST CELL LINE"


# ---- methodology (backend ke bilkul same) ----
# Production day = sir ki query jaisa: 00:30 se agle din 00:30 (30 min peeche karke date).
def prod_day(t):
    return (t - timedelta(minutes=30)).date()


def shift_of(t):
    tod = t.time()
    if tod >= __import__("datetime").time(6, 0) and tod < __import__("datetime").time(14, 30):
        return "A"
    if tod >= __import__("datetime").time(14, 30) and tod < __import__("datetime").time(22, 30):
        return "B"
    return "C"


def first_appearance(rows, key_idx, created_idx, want_ws=None, want_loc=None, ws_idx=None, loc_idx=None):
    """MIN(CREATEDON) per serial (optionally filtered by workstation/location)."""
    firsts = {}
    for r in rows:
        if want_ws is not None and r[ws_idx] != want_ws:
            continue
        if want_loc is not None and r[loc_idx] != want_loc:
            continue
        serial, t = r[key_idx], r[created_idx]
        if serial is None or t is None:
            continue
        # per-workstation distinctness: key = (ws, serial); per-location: serial only
        k = (r[ws_idx], serial) if want_ws is not None else serial
        if k not in firsts or t < firsts[k]:
            firsts[k] = t
    return list(firsts.values())


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["HISTORY"]
    header = [c.value for c in ws[1]]
    i_ws, i_loc, i_serial, i_on = (header.index(x) for x in
                                   ("WORKSTATION", "LOCATION", "SERIALNO", "CREATEDON"))
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    old_firsts = first_appearance(rows, i_serial, i_on, want_ws=OLD_LINE, ws_idx=i_ws)
    new_firsts = first_appearance(rows, i_serial, i_on, want_ws=NEW_LINE, ws_idx=i_ws)
    tc_firsts = first_appearance(rows, i_serial, i_on, want_loc=TEST_CELL, ws_idx=i_ws, loc_idx=i_loc)

    # Saare activity-wale din (sirf new-engine wale nahi) - backend ke availableDays jaisa.
    days = sorted({prod_day(r[i_on]) for r in rows if r[i_on] is not None})

    fails = 0
    checks = 0
    for d in days:
        ds = json.dumps  # noqa
        with urllib.request.urlopen(f"{API}?date={d.isoformat()}", timeout=10) as resp:
            api_data = json.load(resp)

        lo = ds_start = d  # readability
        in_day = lambda t: prod_day(t) == d  # noqa

        # ---- per-day KPI ----
        exp_old = sum(1 for t in old_firsts if in_day(t))
        exp_new = sum(1 for t in new_firsts if in_day(t))
        exp_tc = sum(1 for t in tc_firsts if in_day(t))
        for label, exp, got in (("KPI oldLine", exp_old, api_data["kpis"]["oldLine"]),
                                 ("KPI newLine", exp_new, api_data["kpis"]["newLine"]),
                                 ("KPI testCell", exp_tc, api_data["kpis"]["testCell"])):
            checks += 1
            if exp != got:
                fails += 1
                print(f"  [FAIL] {d} {label}: excel={exp} api={got}")

        # ---- per-shift ----
        for s in ("A", "B", "C"):
            so = sum(1 for t in old_firsts if in_day(t) and shift_of(t) == s)
            sn = sum(1 for t in new_firsts if in_day(t) and shift_of(t) == s)
            stc = sum(1 for t in tc_firsts if in_day(t) and shift_of(t) == s)
            blk = api_data["shifts"][s.lower()]
            for label, exp, got in ((f"Shift {s} oldLine", so, blk["oldLine"]),
                                    (f"Shift {s} newLine", sn, blk["newLine"]),
                                    (f"Shift {s} testCell", stc, blk["testCell"])):
                checks += 1
                if exp != got:
                    fails += 1
                    print(f"  [FAIL] {d} {label}: excel={exp} api={got}")

        # ---- per-hour (old, new, test) - sirf non-zero hours compare (API ab 24 ghante bhejta) ----
        exp_hours = defaultdict(lambda: [0, 0, 0])  # hour -> [old, new, test]
        for t in old_firsts:
            if in_day(t):
                exp_hours[t.hour][0] += 1
        for t in new_firsts:
            if in_day(t):
                exp_hours[t.hour][1] += 1
        for t in tc_firsts:
            if in_day(t):
                exp_hours[t.hour][2] += 1
        exp_hours = {h: v for h, v in exp_hours.items() if any(v)}
        api_hours = {int(x["hour"][:2]): [x["oldLine"], x["newLine"], x["testCell"]]
                     for x in api_data["hourly"] if x["oldLine"] or x["newLine"] or x["testCell"]}
        checks += 1
        if exp_hours != api_hours:
            fails += 1
            print(f"  [FAIL] {d} hourly mismatch:\n     excel={dict(sorted(exp_hours.items()))}\n     api  ={dict(sorted(api_hours.items()))}")

        tot_new = sum(api_data['shifts'][s]['newLine'] for s in 'abc')
        print(f"  {d}  Old={exp_old:>4} New={exp_new:>4} Test={exp_tc:>4}  | A/B/C new={api_data['shifts']['a']['newLine']}/{api_data['shifts']['b']['newLine']}/{api_data['shifts']['c']['newLine']} (sum={tot_new})  OK")

    # ---- Daily + Monthly charts (ek response se - sabme same aata hain) ----
    with urllib.request.urlopen(f"{API}?date={days[-1].isoformat()}", timeout=10) as resp:
        api_data = json.load(resp)

    # Daily
    exp_daily = {}
    for d in days:
        exp_daily[d.strftime("%d %b")] = [
            sum(1 for t in old_firsts if prod_day(t) == d),
            sum(1 for t in new_firsts if prod_day(t) == d),
            sum(1 for t in tc_firsts if prod_day(t) == d),
        ]
    api_daily = {x["date"]: [x["oldLine"], x["newLine"], x["testCell"]] for x in api_data["daily"]}
    checks += 1
    if exp_daily != api_daily:
        fails += 1
        print(f"  [FAIL] daily mismatch:\n     excel={exp_daily}\n     api  ={api_daily}")
    else:
        print(f"  Daily chart: {len(api_daily)} days  OK")

    # Monthly
    exp_monthly = defaultdict(lambda: [0, 0, 0])
    for t in old_firsts:
        exp_monthly[prod_day(t).strftime("%b %Y")][0] += 1
    for t in new_firsts:
        exp_monthly[prod_day(t).strftime("%b %Y")][1] += 1
    for t in tc_firsts:
        exp_monthly[prod_day(t).strftime("%b %Y")][2] += 1
    exp_monthly = dict(exp_monthly)
    api_monthly = {x["month"]: [x["oldLine"], x["newLine"], x["testCell"]] for x in api_data["monthly"]}
    checks += 1
    if exp_monthly != api_monthly:
        fails += 1
        print(f"  [FAIL] monthly mismatch:\n     excel={exp_monthly}\n     api  ={api_monthly}")
    else:
        print(f"  Monthly chart: {len(api_monthly)} months  OK  {api_monthly}")

    print(f"\n==== {checks} checks across {len(days)} days | {fails} FAIL ====")
    print("ALL PASS - dashboard exactly matches the source data." if fails == 0
          else f"{fails} mismatches - fix needed.")


if __name__ == "__main__":
    main()
