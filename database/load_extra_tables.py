"""
#############################################################################
# ⚠️  LOCAL TEST DB ONLY — PRODUCTION / REAL COMPANY DB PE KABHI MAT CHALAO! #
# Ye script in tables ko pehle DROP karta hain phir slim sample bana deta:   #
#   MPI_COB_T_AMI_CAPTURE_LOG, MPI_COB_T_SERIAL_NO, MPI_COB_T_TRANSACTION_OUTBOUND#
# Real DB pe in tables mein CRORES rows hain - chalaya to wo UD jaayega.     #
# Real DB pe ye tables already maujood hain; sirf connection string point    #
# karna hain, kuch load NAHI karna. Ye sirf hamare local FlexNet DB ke liye. #
#############################################################################
"""
import os
import subprocess
import openpyxl

DB = "FlexNet"  # SAFETY: sirf local test DB. Real DB ka naam yahan kabhi mat daalo.
DL = r"C:\Users\Sarfaraz\Downloads"
OUT = os.path.join(os.path.dirname(__file__), "_extra_tables.sql")

# (table, source_xlsx, sheet, [(col, sqltype, kind)])  kind: s=string, n=number, d=datetime
TABLES = [
    ("MPI_COB_T_AMI_CAPTURE_LOG", "AMIUPLOAD.xlsx", "AMIUPLOAD", [
        ("WORKSTATION", "NVARCHAR(40)", "s"),
        ("SERIALNO", "NVARCHAR(40)", "s"),
        ("CREATEDON", "DATETIME2", "d"),
    ]),
    ("MPI_COB_T_SERIAL_NO", "cob_t_serial_no.xlsx", "cob_t_serial_no", [
        ("SERIALNO", "NVARCHAR(40)", "s"),
        ("WORKORDERNO", "NVARCHAR(40)", "s"),
        ("STATUS", "FLOAT", "n"),
        ("CREATEDON", "DATETIME2", "d"),
    ]),
    ("MPI_COB_T_TRANSACTION_OUTBOUND", "COB_T_TRANSACTION_OUTBOUND.xlsx", "COB_T_TRANSACTION_OUTBOUND", [
        ("WIPJOBNO", "NVARCHAR(40)", "s"),
        ("SERIALNO", "NVARCHAR(40)", "s"),
        ("OVERALLSTATUS", "FLOAT", "n"),
        ("CREATEDON", "DATETIME2", "d"),
    ]),
]
BATCH = 1000


def fmt(v, kind):
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return "NULL"
    if kind == "s":
        return "N'" + str(v).strip().replace("'", "''") + "'"
    if kind == "n":
        f = float(v)
        return str(int(f)) if f.is_integer() else repr(f)
    return "'" + v.strftime("%Y-%m-%d %H:%M:%S") + "'"  # datetime


def main():
    # SAFETY GUARD: real/production DB pe galti se chal gaya to tables UD jaate.
    if DB.strip().lower() != "flexnet":
        raise SystemExit(
            f"REFUSING TO RUN: DB='{DB}'. Ye script sirf local 'FlexNet' test DB ke liye hain. "
            "Real DB pe chalana = tables DROP. Connection string sirf app ke liye badlo, ye script nahi.")

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("SET NOCOUNT ON;\n\n")
        for table, xlsx, sheet, cols in TABLES:
            print(f"Reading {xlsx} ...")
            wb = openpyxl.load_workbook(os.path.join(DL, xlsx), data_only=True)
            ws = wb[sheet]
            header = [str(c.value) for c in ws[1]]
            idx = [header.index(c) for c, _, _ in cols]
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            col_defs = ", ".join(f"{c} {t}" for c, t, _ in cols)
            col_list = ", ".join(c for c, _, _ in cols)
            f.write(f"IF OBJECT_ID('dbo.{table}','U') IS NOT NULL DROP TABLE dbo.{table};\nGO\n")
            f.write(f"CREATE TABLE dbo.{table} ({col_defs});\nGO\n")

            kinds = [k for _, _, k in cols]
            for i in range(0, len(rows), BATCH):
                batch = rows[i:i + BATCH]
                f.write(f"INSERT INTO dbo.{table} ({col_list}) VALUES\n")
                vals = []
                for r in batch:
                    cells = [fmt(r[idx[j]] if idx[j] < len(r) else None, kinds[j]) for j in range(len(cols))]
                    vals.append("(" + ", ".join(cells) + ")")
                f.write(",\n".join(vals) + ";\nGO\n")
            print(f"  {table}: {len(rows)} rows")

    print("Running sqlcmd ...")
    res = subprocess.run(
        ["sqlcmd", "-S", "localhost", "-d", DB, "-E", "-C", "-i", OUT],
        capture_output=True, text=True)
    print(res.stdout[-500:] if res.stdout else "")
    if res.returncode != 0:
        print("ERROR:", res.stderr[-1000:])
    else:
        print("Done - tables loaded.")


if __name__ == "__main__":
    main()
